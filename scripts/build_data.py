from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw"
OUT_DIR = ROOT / "data" / "processed"
TARGET_SHEET = "2C炼厂加工进口原油详解"
INCLUDE_TOTAL_ROWS = False

LONG_COLUMNS = [
    "month",
    "year",
    "region",
    "refinery",
    "oil_name_en",
    "oil_name_cn",
    "origin_region",
    "volume_10kt",
    "source_file",
    "source_sheet",
    "source_row",
    "source_col",
    "is_total_row",
]

LOG_COLUMNS = [
    "source_file",
    "month",
    "status",
    "sheet_found",
    "raw_data_rows",
    "records_generated",
    "empty_oil_name_count",
    "used_cn_as_oil_name_count",
    "empty_refinery_count",
    "invalid_number_count",
    "source_total_10kt",
    "detail_total_10kt",
    "total_diff_10kt",
    "total_diff_pct",
    "messages",
]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def has_cjk(value: str) -> bool:
    return any("\u4e00" <= char <= "\u9fff" for char in value)


def remove_cjk(value: str) -> str:
    return "".join(char for char in value if not ("\u4e00" <= char <= "\u9fff")).strip()


def parse_month_from_name(filename: str) -> tuple[str | None, str | None]:
    matches = list(re.finditer(r"(?<!\d)(?:20)?(?P<yy>\d{2})(?P<mm>0[1-9]|1[0-2])(?!\d)", filename))
    if not matches:
        return None, None
    match = matches[-1]
    year = f"20{match.group('yy')}"
    return f"{year}-{match.group('mm')}", year


def to_number(value: Any) -> tuple[float | None, bool]:
    if value is None:
        return None, False
    if isinstance(value, str) and value.strip() in {"", "-", "--", "/"}:
        return None, False
    if isinstance(value, (int, float)):
        return float(value), False
    try:
        return float(str(value).replace(",", "").strip()), False
    except ValueError:
        return None, True


def find_target_sheet(workbook) -> str | None:
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip() == TARGET_SHEET:
            return sheet_name
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().startswith("2C") and "进口原油" in sheet_name:
            return sheet_name
    return None


def column_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def parse_workbook(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    month, year = parse_month_from_name(path.name)
    log = {key: "" for key in LOG_COLUMNS}
    log.update(
        {
            "source_file": path.name,
            "month": month or "",
            "status": "ok",
            "sheet_found": False,
            "raw_data_rows": 0,
            "records_generated": 0,
            "empty_oil_name_count": 0,
            "used_cn_as_oil_name_count": 0,
            "empty_refinery_count": 0,
            "invalid_number_count": 0,
            "source_total_10kt": 0,
            "detail_total_10kt": 0,
            "total_diff_10kt": 0,
            "total_diff_pct": 0,
            "messages": "",
        }
    )
    messages: list[str] = []

    if not month or not year:
        log["status"] = "skipped"
        log["messages"] = "未能从文件名识别月份"
        return [], log

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        log["status"] = "error"
        log["messages"] = f"打开文件失败: {exc}"
        return [], log

    sheet_name = find_target_sheet(wb)
    print(f"{path.name}: month={month}, sheet_found={bool(sheet_name)}")
    if not sheet_name:
        log["status"] = "skipped"
        log["messages"] = f"缺失sheet: {TARGET_SHEET}"
        return [], log

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 5:
        log["status"] = "skipped"
        log["sheet_found"] = True
        log["messages"] = "sheet行数不足，无法解析"
        return [], log

    log["sheet_found"] = True
    header_origin = rows[1]
    header_cn = rows[2]
    header_en = rows[3]
    max_col = max(len(header_origin), len(header_cn), len(header_en))

    origin_by_col: dict[int, str] = {}
    current_origin = ""
    for col in range(5, max_col + 1):
        origin = normalize_text(header_origin[col - 1] if len(header_origin) >= col else "")
        if origin:
            current_origin = origin
        origin_by_col[col] = current_origin

    records: list[dict[str, Any]] = []
    current_region = ""
    source_total = 0.0
    detail_total = 0.0

    for row_index, row in enumerate(rows[4:], start=5):
        region_cell = normalize_text(row[1] if len(row) >= 2 else "")
        refinery = normalize_text(row[2] if len(row) >= 3 else "")
        total_value, total_invalid = to_number(row[3] if len(row) >= 4 else None)
        if region_cell:
            current_region = region_cell
        if not refinery:
            log["empty_refinery_count"] += 1
            continue
        is_total_row = "合计" in refinery
        if is_total_row and not INCLUDE_TOTAL_ROWS:
            continue

        log["raw_data_rows"] += 1
        if total_invalid:
            log["invalid_number_count"] += 1
            messages.append(f"第{row_index}行D列合计无法转数字")
        elif total_value:
            source_total += total_value

        for col in range(5, max_col + 1):
            value = row[col - 1] if len(row) >= col else None
            volume, invalid = to_number(value)
            if invalid:
                log["invalid_number_count"] += 1
                messages.append(f"第{row_index}行{column_letter(col)}列数量无法转数字: {value}")
                continue
            if volume is None or volume == 0:
                continue

            oil_name_cn = normalize_text(header_cn[col - 1] if len(header_cn) >= col else "")
            oil_name_en = normalize_text(header_en[col - 1] if len(header_en) >= col else "")
            if oil_name_en == "沙特原油" and oil_name_cn.upper() == "ARABIAN":
                oil_name_en, oil_name_cn = "ARABIAN", "沙特原油"
                messages.append(f"{column_letter(col)}列中英文油种名反置，已修正为 ARABIAN / 沙特原油")
            if has_cjk(oil_name_en) and any(char.isalpha() for char in oil_name_en):
                cleaned_oil_name_en = remove_cjk(oil_name_en)
                if cleaned_oil_name_en:
                    messages.append(f"{column_letter(col)}列英文油种名含中文字符，已清洗为: {cleaned_oil_name_en}")
                    oil_name_en = cleaned_oil_name_en
            if not oil_name_en:
                log["empty_oil_name_count"] += 1
                if oil_name_cn:
                    oil_name_en = oil_name_cn
                    log["used_cn_as_oil_name_count"] += 1
                    messages.append(f"{column_letter(col)}列英文油种名为空，已使用中文名: {oil_name_cn}")
                else:
                    messages.append(f"{column_letter(col)}列中英文油种名均为空，已跳过")
                    continue

            detail_total += volume
            records.append(
                {
                    "month": month,
                    "year": year,
                    "region": current_region,
                    "refinery": refinery,
                    "oil_name_en": oil_name_en,
                    "oil_name_cn": oil_name_cn,
                    "origin_region": origin_by_col.get(col, ""),
                    "volume_10kt": volume,
                    "source_file": path.name,
                    "source_sheet": TARGET_SHEET,
                    "source_row": row_index,
                    "source_col": col,
                    "is_total_row": is_total_row,
                }
            )

    log["source_total_10kt"] = round(source_total, 6)
    log["detail_total_10kt"] = round(detail_total, 6)
    diff = detail_total - source_total
    log["total_diff_10kt"] = round(diff, 6)
    log["total_diff_pct"] = round(diff / source_total, 6) if source_total else 0
    if source_total and abs(diff) > max(1.0, source_total * 0.01):
        messages.append("明细合计与D列合计差异超过1万吨或1%")

    log["records_generated"] = len(records)
    log["messages"] = "；".join(dict.fromkeys(messages))
    print(f"{path.name}: records={len(records)}, source_total={source_total:.2f}, detail_total={detail_total:.2f}")
    return records, log


def merge_duplicates(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    for record in records:
        key = (
            record["month"],
            record["year"],
            record["region"],
            record["refinery"],
            record["oil_name_en"],
            record["oil_name_cn"],
            record["origin_region"],
            record["source_sheet"],
            record["is_total_row"],
        )
        if key not in grouped:
            grouped[key] = dict(record)
            continue
        grouped[key]["volume_10kt"] += record["volume_10kt"]
        grouped[key]["source_file"] = "|".join(sorted(set(str(grouped[key]["source_file"]).split("|") + [record["source_file"]])))
        grouped[key]["source_row"] = f"{grouped[key]['source_row']}|{record['source_row']}"
        grouped[key]["source_col"] = f"{grouped[key]['source_col']}|{record['source_col']}"

    merged = list(grouped.values())
    for record in merged:
        record["volume_10kt"] = round(float(record["volume_10kt"]), 6)
    return sorted(merged, key=lambda item: (item["month"], item["refinery"], item["oil_name_en"]))


def write_outputs(records: list[dict[str, Any]], logs: list[dict[str, Any]]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    csv_path = OUT_DIR / "refinery_import_crude_long.csv"
    json_path = OUT_DIR / "refinery_import_crude_long.json"
    log_path = OUT_DIR / "processing_log.csv"

    df = pd.DataFrame(records, columns=LONG_COLUMNS)
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    payload = {
        "schema_version": "1.0",
        "unit": "万吨",
        "generated_from": TARGET_SHEET,
        "record_count": len(records),
        "months": sorted(df["month"].dropna().unique().tolist()) if not df.empty else [],
        "records": records,
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    with log_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=LOG_COLUMNS)
        writer.writeheader()
        writer.writerows(logs)

    print(f"wrote {csv_path}")
    print(f"wrote {json_path}")
    print(f"wrote {log_path}")


def main() -> None:
    files = sorted([*RAW_DIR.glob("*.xlsx"), *RAW_DIR.glob("*.xlsm")])
    if not files:
        raise SystemExit(f"未找到Excel文件，请放入: {RAW_DIR}")

    all_records: list[dict[str, Any]] = []
    logs: list[dict[str, Any]] = []
    for path in files:
        records, log = parse_workbook(path)
        all_records.extend(records)
        logs.append(log)

    merged_records = merge_duplicates(all_records)
    duplicate_count = len(all_records) - len(merged_records)
    if duplicate_count:
        logs.append(
            {
                **{key: "" for key in LOG_COLUMNS},
                "source_file": "__all__",
                "status": "ok",
                "records_generated": len(merged_records),
                "messages": f"合并重复记录 {duplicate_count} 条",
            }
        )
    write_outputs(merged_records, logs)
    print(f"total_records={len(merged_records)}")


if __name__ == "__main__":
    main()
